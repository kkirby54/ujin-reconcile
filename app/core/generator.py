"""
엑셀 파일 생성 모듈
Upload_form.xlsx 형식의 전표 파일 생성
"""
import io
from openpyxl import load_workbook


# Unmatched 시트 헤더
HEADERS_UNMATCHED = [
    "거래일자", "입출금구분", "금액", "보낸분/받는분", "메모", "사유"
]

# Upload_form.xlsx의 컬럼 순서 (12행 헤더 기준)
UPLOAD_FORM_COLUMNS = [
    '순번', '거래일', '전표구분', '코드', '거래처명', '적요', '결제장부', '금액',
    '차변계정코드', '차변금액', '대변계정코드', '대변금액', '메모', '프로젝트', '은행코드'
]


def build_upload_form_workbook(matches, unmatched, template_path="Upload_form.xlsx"):
    """
    매칭된 전표 데이터를 Upload_form.xlsx 형식으로 직접 생성
    
    Args:
        matches (list): 매칭된 거래 데이터 리스트
        unmatched (list): 매칭 실패한 거래 데이터 리스트
        template_path (str): Upload_form.xlsx 템플릿 파일 경로
        
    Returns:
        io.BytesIO: Upload_form.xlsx 형식의 엑셀 파일 바이너리 데이터
    """
    # Upload_form.xlsx 템플릿 로드 (서식 보존)
    template_wb = load_workbook(template_path)
    template_ws = template_wb['일반전표']
    
    # 데이터 입력 시작 행 (12행이 헤더이므로 13행부터)
    start_row = 13
    
    # 매칭된 거래 데이터를 Upload_form 형식으로 변환하여 입력
    sequence_number = 1
    for row_idx, match_row in enumerate(matches, start=start_row):
        # 전표구분: 입금=2, 출금=1
        voucher_type = 2 if match_row["type"] == "입금" else 1
        
        # 금액 설정
        amount = match_row["amount"]
        
        # 차변/대변 금액 설정
        debit_amount = match_row["amount"] if match_row["type"] == "입금" else ""
        credit_amount = match_row["amount"] if match_row["type"] == "출금" else ""
        
        # Upload_form 컬럼 순서에 맞춰 데이터 입력
        row_data = [
            sequence_number,              # 순번
            match_row["date"],            # 거래일
            voucher_type,                 # 전표구분
            match_row["code"],            # 코드
            match_row["name"],            # 거래처명
            match_row["memo"],            # 적요
            2,                            # 결제장부 (예금 고정)
            amount,                       # 금액
            "",                           # 차변계정코드
            debit_amount,                 # 차변금액
            "",                           # 대변계정코드
            credit_amount,                # 대변금액
            match_row["memo"],            # 메모
            "",                           # 프로젝트
            ""                            # 은행코드
        ]
        
        # 각 셀에 값 입력 (서식 보존)
        for col_idx, value in enumerate(row_data, start=1):
            template_ws.cell(row=row_idx, column=col_idx, value=value)
        
        sequence_number += 1
    
    # Unmatched 시트가 있으면 데이터 추가, 없으면 생성
    if "Unmatched" in template_wb.sheetnames:
        ws_unmatched = template_wb["Unmatched"]
        # 기존 데이터 삭제 (헤더 제외)
        ws_unmatched.delete_rows(2, ws_unmatched.max_row)
    else:
        ws_unmatched = template_wb.create_sheet("Unmatched")
        ws_unmatched.append(HEADERS_UNMATCHED)
    
    # 매칭 실패한 거래 데이터 추가
    for unmatched_row in unmatched:
        ws_unmatched.append([
            unmatched_row["date"],           # 거래일자
            unmatched_row["type"],           # 입출금구분
            unmatched_row["amount"],         # 금액
            unmatched_row["counter_raw"],    # 보낸분/받는분
            unmatched_row["memo"],           # 메모
            "매칭 실패"                      # 사유
        ])
    
    # 바이너리 데이터로 변환
    buffer = io.BytesIO()
    template_wb.save(buffer)
    buffer.seek(0)
    
    return buffer
